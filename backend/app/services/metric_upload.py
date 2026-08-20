import csv
import io

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models.collaboration import Collaboration
from app.db.models.enums import CollabStage, MetricImportStatus
from app.db.models.metric_import import MetricImport
from app.schemas.metric_upload import MetricUploadResult

# Exact header names the upload template uses -- validated before any row is
# processed, per "Column names are validated before any record changes."
REQUIRED_COLUMNS = ["POC Code", "Video Link", "Views", "Likes", "Comments", "Revenue", "Ad Spend", "ROAS"]


def _parse_rows(filename: str, raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Same CSV/Excel-agnostic parsing as bulk_upload_creators in
    creators.py, adapted to also return the header row so it can be
    validated even when the sheet has zero data rows."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig")))
        header = reader.fieldnames or []
        return list(header), list(reader)

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    sheet_rows = sheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(sheet_rows, [])]
    rows = [
        {header[i]: ("" if cell is None else str(cell)) for i, cell in enumerate(row) if i < len(header)}
        for row in sheet_rows
    ]
    return header, rows


def _parse_int(raw: str | None) -> int | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return int(float(raw))  # tolerate Excel's "1234.0"-style cell stringification
    except ValueError:
        return None


def _parse_decimal(raw: str | None) -> float | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


async def _record_import(
    db: AsyncSession,
    filename: str,
    total_rows: int,
    updated_rows: int,
    skipped_rows: int,
    status: MetricImportStatus,
    uploaded_by: int,
) -> None:
    db.add(
        MetricImport(
            filename=filename,
            total_rows=total_rows,
            updated_rows=updated_rows,
            skipped_rows=skipped_rows,
            status=status,
            uploaded_by=uploaded_by,
        )
    )
    await db.commit()


async def process_metric_upload(
    db: AsyncSession, filename: str, raw: bytes, uploaded_by: int
) -> MetricUploadResult:
    lower = filename.lower()
    if not (lower.endswith(".csv") or lower.endswith(".xlsx") or lower.endswith(".xls")):
        await _record_import(db, filename, 0, 0, 0, MetricImportStatus.failed, uploaded_by)
        return MetricUploadResult(total_rows=0, updated=0, skipped=0, errors=["Upload a .csv, .xlsx or .xls file"])

    try:
        header, rows = _parse_rows(filename, raw)
    except Exception:
        await _record_import(db, filename, 0, 0, 0, MetricImportStatus.failed, uploaded_by)
        return MetricUploadResult(
            total_rows=0, updated=0, skipped=0, errors=["Could not read this file. Check the format and try again."]
        )

    missing_columns = [col for col in REQUIRED_COLUMNS if col not in header]
    if missing_columns:
        await _record_import(db, filename, 0, 0, 0, MetricImportStatus.failed, uploaded_by)
        return MetricUploadResult(
            total_rows=0, updated=0, skipped=0, errors=[f"Missing required column(s): {', '.join(missing_columns)}"]
        )

    updated = 0
    skipped = 0
    errors: list[str] = []

    for line_num, row in enumerate(rows, start=2):
        poc_code = (row.get("POC Code") or "").strip()
        if not poc_code:
            errors.append(f"Row {line_num}: missing POC code")
            skipped += 1
            continue

        collab = (
            await db.execute(select(Collaboration).where(Collaboration.poc_code == poc_code))
        ).scalar_one_or_none()
        if collab is None:
            errors.append(f"Row {line_num}: no collaboration found for POC code '{poc_code}'")
            skipped += 1
            continue
        if collab.stage != CollabStage.live:
            errors.append(f"Row {line_num}: '{poc_code}' is not a Live record")
            skipped += 1
            continue

        video_link = (row.get("Video Link") or "").strip()
        if not video_link:
            errors.append(f"Row {line_num}: missing Video Link")
            skipped += 1
            continue
        existing_link = (collab.video_link or "").strip().lower()
        if video_link.lower() != existing_link:
            errors.append(f"Row {line_num}: video link does not match the Live record for '{poc_code}'")
            skipped += 1
            continue

        views = _parse_int(row.get("Views"))
        likes = _parse_int(row.get("Likes"))
        comments = _parse_int(row.get("Comments"))
        revenue = _parse_decimal(row.get("Revenue"))
        ad_spend = _parse_decimal(row.get("Ad Spend"))
        roas = _parse_decimal(row.get("ROAS"))

        # A blank cell leaves the existing field unchanged -- a later, partial
        # weekly sheet must never null out previously-uploaded values.
        if views is not None:
            collab.views_count = views
        if likes is not None:
            collab.likes_count = likes
        if comments is not None:
            collab.comments_count = comments
        if revenue is not None:
            collab.revenue = revenue
        if ad_spend is not None:
            collab.ad_spend = ad_spend
        if roas is not None:
            collab.roas = roas
        elif collab.revenue is not None and collab.ad_spend not in (None, 0):
            # Fallback computed from whatever now sits on the record (this
            # row's values, already applied above, or previously-stored ones).
            collab.roas = round(float(collab.revenue) / float(collab.ad_spend), 2)

        updated += 1

    total_rows = len(rows)
    await _record_import(db, filename, total_rows, updated, skipped, MetricImportStatus.completed, uploaded_by)
    return MetricUploadResult(total_rows=total_rows, updated=updated, skipped=skipped, errors=errors)
