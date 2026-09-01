import csv
import io
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.deps import (
    get_current_user,
    owner_scope_filter,
    require_admin,
    require_creator_workspace_access,
    scoped_owner_ids,
)
from app.db.models.collaboration import Collaboration
from app.db.models.collaboration_product import CollaborationProduct
from app.db.models.creator import Creator
from app.db.models.creator_file import CreatorFile
from app.db.models.enums import CollabStage, CreatorStage, FollowUpStatus, OwnershipEventType, UserRole
from app.db.models.follow_up import FollowUp
from app.db.models.message import Message
from app.db.models.ownership_event import OwnershipEvent
from app.db.models.product import Product
from app.db.models.stage_event import StageEvent
from app.db.models.user import User
from app.db.session import get_db
from app.schemas.creator import (
    BoardStats,
    BulkAssign,
    BulkUploadResult,
    BulkUploadRowResult,
    CreatorCreate,
    CreatorDetail,
    CreatorListResponse,
    CreatorOut,
    CreatorTableResponse,
    CreatorTableRow,
    CreatorUpdate,
    NextAction,
    OwnershipMatch,
    StageTimelineEntry,
    StageTransition,
    TransferOwnership,
)
from app.schemas.creator_file import CreatorFileOut
from app.schemas.creator_lifecycle import CreatorLifecycle
from app.schemas.message import MessageCreate, MessageOut
from app.services.collab_pipeline import COLLAB_STAGE_INDEX, COLLAB_STAGE_LABELS, effective_live_dates, is_video_live
from app.services.creator_lifecycle import get_creator_lifecycle
from app.services.pipeline import STAGE_INDEX, STAGE_LABELS, STAGE_ORDER

router = APIRouter(
    prefix="/creators", tags=["creators"], dependencies=[Depends(require_creator_workspace_access)]
)

UPLOAD_ROOT = Path(settings.upload_dir)


async def _get_creator_or_404(creator_id: int, db: AsyncSession, user: User | None = None) -> Creator:
    creator = await db.get(Creator, creator_id)
    if creator is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Creator not found")
    if user is not None:
        owner_ids = await scoped_owner_ids(user, db)
        if owner_ids is not None and creator.owner_id not in owner_ids:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Creator not found")
        # Archived leads are visible in the Database list to everyone, but
        # opening one (detail, lifecycle, messages, edits, ...) is admin-only.
        if creator.is_archived and user.role != UserRole.admin:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="This lead is archived -- only an admin can open it.",
            )
    return creator


@router.get("/check-ownership", response_model=list[OwnershipMatch])
async def check_ownership(
    query: str = Query(..., min_length=2),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
) -> list[Creator]:
    pattern = f"%{query}%"
    result = await db.execute(
        select(Creator).where(
            or_(Creator.instagram_handle.ilike(pattern), Creator.phone.ilike(pattern))
        )
    )
    return list(result.scalars().all())


@router.get("/board-stats", response_model=BoardStats)
async def board_stats(
    owner_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> BoardStats:
    owner_ids = await owner_scope_filter(user, db, owner_id)
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    week_start = today_start - timedelta(days=today_start.weekday())

    active_creators_stmt = select(func.count(Creator.id)).where(Creator.current_stage != CreatorStage.paid)
    overdue_actions_stmt = select(func.count(FollowUp.id)).where(
        FollowUp.status == FollowUpStatus.open, FollowUp.due_at < now
    )
    due_today_stmt = select(func.count(FollowUp.id)).where(
        FollowUp.status == FollowUpStatus.open, FollowUp.due_at >= today_start, FollowUp.due_at < today_end
    )
    completed_today_stmt = select(func.count(FollowUp.id)).where(
        FollowUp.status == FollowUpStatus.done,
        FollowUp.completed_at >= today_start,
        FollowUp.completed_at < today_end,
    )
    due_this_week_stmt = select(func.count(FollowUp.id)).where(FollowUp.due_at >= week_start)
    completed_this_week_stmt = select(func.count(FollowUp.id)).where(
        FollowUp.status == FollowUpStatus.done, FollowUp.completed_at >= week_start
    )

    if owner_ids is not None:
        active_creators_stmt = active_creators_stmt.where(Creator.owner_id.in_(owner_ids))
        overdue_actions_stmt = overdue_actions_stmt.where(FollowUp.assigned_to.in_(owner_ids))
        due_today_stmt = due_today_stmt.where(FollowUp.assigned_to.in_(owner_ids))
        completed_today_stmt = completed_today_stmt.where(FollowUp.assigned_to.in_(owner_ids))
        due_this_week_stmt = due_this_week_stmt.where(FollowUp.assigned_to.in_(owner_ids))
        completed_this_week_stmt = completed_this_week_stmt.where(FollowUp.assigned_to.in_(owner_ids))

    active_creators = (await db.execute(active_creators_stmt)).scalar_one()
    overdue_actions = (await db.execute(overdue_actions_stmt)).scalar_one()
    due_today = (await db.execute(due_today_stmt)).scalar_one()
    completed_today = (await db.execute(completed_today_stmt)).scalar_one()
    due_this_week = (await db.execute(due_this_week_stmt)).scalar_one()
    completed_this_week = (await db.execute(completed_this_week_stmt)).scalar_one()
    target_pct = (completed_this_week / due_this_week * 100) if due_this_week else 100.0

    return BoardStats(
        active_creators=active_creators,
        overdue_actions=overdue_actions,
        due_today=due_today,
        weekly_target_pct=round(target_pct, 0),
        follow_ups_completed_today=completed_today,
        follow_ups_total_today=due_today + completed_today,
    )


@router.get("", response_model=CreatorListResponse)
async def list_creators(
    owner_id: int | None = None,
    stage: CreatorStage | None = None,
    category: str | None = None,
    search: str | None = None,
    limit: int = Query(20, le=500),
    offset: int = 0,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> CreatorListResponse:
    owner_ids = await owner_scope_filter(user, db, owner_id)

    # Archived creators can't be picked for a new collaboration — they'd need
    # to be unarchived first, which only an admin can see to do.
    stmt = select(Creator).where(Creator.is_archived.is_(False))
    if owner_ids is not None:
        stmt = stmt.where(Creator.owner_id.in_(owner_ids))
    if stage is not None:
        stmt = stmt.where(Creator.current_stage == stage)
    if category is not None:
        stmt = stmt.where(Creator.category == category)
    if search:
        pattern = f"%{search}%"
        stmt = stmt.where(
            or_(Creator.name.ilike(pattern), Creator.instagram_handle.ilike(pattern), Creator.phone.ilike(pattern))
        )

    total = (await db.execute(select(func.count()).select_from(stmt.subquery()))).scalar_one()
    result = await db.execute(stmt.order_by(Creator.last_activity_at.desc()).limit(limit).offset(offset))
    items = list(result.scalars().all())
    return CreatorListResponse(items=items, total=total)


SORTABLE_FIELDS = {
    "name",
    "followers_count",
    "videos_delivered",
    "last_cost",
    "last_video_live_at",
    "comments_count",
    "created_at",
    "current_stage",
    "archived_at",
}


@router.get("/table", response_model=CreatorTableResponse)
async def list_creators_table(
    owner_id: int | None = None,
    search: str | None = None,
    is_archived: bool = False,
    sort_by: str = Query("created_at"),
    sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
    limit: int = Query(10, le=5000),
    offset: int = 0,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> CreatorTableResponse:
    if sort_by not in SORTABLE_FIELDS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid sort_by")
    owner_ids = await owner_scope_filter(user, db, owner_id)
    # Archived leads are visible in this list to every role (scoped to their
    # normal owner_ids like any other row here) -- opening one is what's
    # admin-gated, enforced separately in _get_creator_or_404.
    stmt = select(Creator).where(Creator.is_archived == is_archived)
    if owner_ids is not None:
        stmt = stmt.where(Creator.owner_id.in_(owner_ids))
    if search:
        pattern = f"%{search}%"
        stmt = stmt.where(
            or_(Creator.name.ilike(pattern), Creator.instagram_handle.ilike(pattern), Creator.phone.ilike(pattern))
        )
    creators = list((await db.execute(stmt)).scalars().all())
    creator_ids = [c.id for c in creators]

    collabs_by_creator: dict[int, list[tuple[Collaboration, str]]] = {cid: [] for cid in creator_ids}
    if creator_ids:
        collab_result = await db.execute(
            select(Collaboration, Product.name)
            .join(CollaborationProduct, CollaborationProduct.collaboration_id == Collaboration.id)
            .join(Product, Product.id == CollaborationProduct.product_id)
            .where(Collaboration.creator_id.in_(creator_ids), CollaborationProduct.is_primary.is_(True))
        )
        for collab, product_name in collab_result.all():
            collabs_by_creator[collab.creator_id].append((collab, product_name))

    # "Last video live" per creator: the most recent (by effective live
    # date -- explicit video_live_date if set, else first transition to
    # Live, see collab_pipeline.effective_live_dates) of their primary-
    # product Live collaborations, reusing collabs_by_creator above rather
    # than a second query.
    live_collab_ids = [
        c.id for collabs in collabs_by_creator.values() for c, _ in collabs if c.stage == CollabStage.live
    ]
    effective_dates = await effective_live_dates(db, live_collab_ids)
    last_live_by_creator: dict[int, tuple[datetime, str, int | None]] = {}
    for creator_id_, collabs in collabs_by_creator.items():
        candidates = [
            (effective_dates[c.id], c, product_name)
            for c, product_name in collabs
            if c.stage == CollabStage.live and c.id in effective_dates
        ]
        if candidates:
            best_date, best_collab, best_product = max(candidates, key=lambda t: t[0])
            last_live_by_creator[creator_id_] = (
                datetime.combine(best_date, datetime.min.time(), tzinfo=timezone.utc),
                best_product,
                best_collab.comments_count,
            )

    rows: list[CreatorTableRow] = []
    for creator in creators:
        collabs = collabs_by_creator.get(creator.id, [])
        videos_delivered = sum(1 for c, _ in collabs if is_video_live(c.stage))
        last_cost = None
        current_stage_label = None
        if collabs:
            most_recent = max(collabs, key=lambda pair: pair[0].last_activity_at)
            current_stage_label = COLLAB_STAGE_LABELS[most_recent[0].stage]
            with_cost = [c for c, _ in collabs if c.commercial_amount is not None]
            if with_cost:
                latest_with_cost = max(with_cost, key=lambda c: c.last_activity_at)
                last_cost = float(latest_with_cost.commercial_amount)
        last_live_at, last_live_product, last_live_comments = last_live_by_creator.get(creator.id, (None, None, None))

        rows.append(
            CreatorTableRow(
                id=creator.id,
                name=creator.name,
                instagram_handle=creator.instagram_handle,
                phone=creator.phone,
                category=creator.category,
                followers_count=creator.followers_count,
                owner_id=creator.owner_id,
                status=creator.status,
                is_archived=creator.is_archived,
                archived_at=creator.archived_at,
                archive_reason=creator.archive_reason,
                created_at=creator.created_at,
                videos_delivered=videos_delivered,
                last_cost=last_cost,
                last_video_live_at=last_live_at,
                last_video_product_name=last_live_product,
                comments_count=last_live_comments,
                current_collab_stage_label=current_stage_label,
            )
        )

    label_to_stage = {label: stage for stage, label in COLLAB_STAGE_LABELS.items()}
    sort_attr = "current_collab_stage_label" if sort_by == "current_stage" else sort_by

    def sort_value(row: CreatorTableRow):
        value = getattr(row, sort_attr)
        if sort_by == "current_stage":
            stage = label_to_stage.get(value)
            value = COLLAB_STAGE_INDEX[stage] if stage is not None else None
        return value

    with_value = [r for r in rows if sort_value(r) is not None]
    without_value = [r for r in rows if sort_value(r) is None]
    with_value.sort(key=sort_value, reverse=(sort_dir == "desc"))
    rows = with_value + without_value
    total = len(rows)
    page = rows[offset : offset + limit]
    return CreatorTableResponse(items=page, total=total)


@router.post("/bulk-upload", response_model=BulkUploadResult)
async def bulk_upload_creators(
    upload: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> BulkUploadResult:
    filename = (upload.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload a .csv, .xlsx or .xls file")

    raw = await upload.read()
    if filename.endswith(".csv"):
        reader: list[dict[str, str]] = list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
    else:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        sheet_rows = sheet.iter_rows(values_only=True)
        header = [str(cell).strip() if cell is not None else "" for cell in next(sheet_rows, [])]
        reader = [
            {header[i]: ("" if cell is None else str(cell)) for i, cell in enumerate(row) if i < len(header)}
            for row in sheet_rows
        ]

    owner_by_email: dict[str, int] = {}
    if user.role == UserRole.admin:
        result = await db.execute(select(User.email, User.id).where(User.role == UserRole.advisor))
        owner_by_email = {email.lower(): uid for email, uid in result.all()}
    users_by_id = {u.id: u.name for u in (await db.execute(select(User))).scalars().all()}

    created = 0
    skipped = 0
    errors: list[str] = []
    row_results: list[BulkUploadRowResult] = []

    for line_num, row in enumerate(reader, start=2):
        name = (row.get("name") or "").strip()
        handle = (row.get("instagram_handle") or row.get("handle") or "").strip().lstrip("@")
        if not name or not handle:
            reason = "Missing name or instagram_handle"
            errors.append(f"Row {line_num}: {reason.lower()}")
            row_results.append(
                BulkUploadRowResult(
                    row=line_num, instagram_handle=handle, name=name, status="error", reason=reason
                )
            )
            continue

        existing = (
            await db.execute(select(Creator).where(Creator.instagram_handle == handle))
        ).scalar_one_or_none()
        if existing is not None:
            skipped += 1
            row_results.append(
                BulkUploadRowResult(
                    row=line_num,
                    instagram_handle=handle,
                    name=name,
                    status="already_present",
                    reason="Already present in the database",
                    existing_owner=users_by_id.get(existing.owner_id, "Unknown"),
                    existing_stage=existing.current_stage.value.replace("_", " ").title(),
                )
            )
            continue

        owner_id = user.id
        if user.role == UserRole.admin:
            owner_email = (row.get("owner_email") or "").strip().lower()
            if owner_email and owner_email in owner_by_email:
                owner_id = owner_by_email[owner_email]

        followers_raw = (row.get("followers_count") or row.get("followers") or "0").strip()
        try:
            followers_count = int(followers_raw) if followers_raw else 0
        except ValueError:
            followers_count = 0

        creator = Creator(
            name=name,
            instagram_handle=handle,
            phone=(row.get("phone") or "").strip() or None,
            category=(row.get("category") or "Beauty").strip(),
            followers_count=followers_count,
            owner_id=owner_id,
        )
        db.add(creator)
        await db.flush()
        db.add(
            OwnershipEvent(
                creator_id=creator.id,
                user_id=owner_id,
                event_type=OwnershipEventType.assigned,
                actor_id=user.id,
            )
        )
        created += 1
        row_results.append(
            BulkUploadRowResult(
                row=line_num, instagram_handle=handle, name=name, status="created", reason="Added to database"
            )
        )

    await db.commit()
    return BulkUploadResult(created=created, skipped=skipped, errors=errors, rows=row_results)


@router.get("/{creator_id}", response_model=CreatorOut)
async def get_creator(
    creator_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Creator:
    return await _get_creator_or_404(creator_id, db, user)


@router.get("/{creator_id}/detail", response_model=CreatorDetail)
async def get_creator_detail(
    creator_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> CreatorDetail:
    creator = await _get_creator_or_404(creator_id, db, user)

    owner = await db.get(User, creator.owner_id)
    owner_name = owner.name if owner else "Unassigned"

    next_follow_up = (
        await db.execute(
            select(FollowUp)
            .where(FollowUp.creator_id == creator_id, FollowUp.status == FollowUpStatus.open)
            .order_by(FollowUp.due_at)
            .limit(1)
        )
    ).scalar_one_or_none()
    next_action = (
        NextAction(
            id=next_follow_up.id,
            type=next_follow_up.type,
            due_at=next_follow_up.due_at,
            priority=next_follow_up.priority.value,
        )
        if next_follow_up
        else None
    )

    stage_events = (
        await db.execute(
            select(StageEvent).where(StageEvent.creator_id == creator_id).order_by(StageEvent.created_at)
        )
    ).scalars().all()
    event_by_stage = {event.to_stage: event for event in stage_events}

    current_index = STAGE_INDEX[creator.current_stage]
    stage_timeline: list[StageTimelineEntry] = []
    for stage in STAGE_ORDER:
        idx = STAGE_INDEX[stage]
        event = event_by_stage.get(stage)
        stage_timeline.append(
            StageTimelineEntry(
                stage=stage,
                label=STAGE_LABELS[stage],
                status="done" if idx < current_index else ("current" if idx == current_index else "upcoming"),
                event_date=event.created_at if event else None,
                note=event.note if event else None,
            )
        )

    return CreatorDetail(
        creator=creator,
        owner_name=owner_name,
        next_action=next_action,
        stage_timeline=stage_timeline,
    )


@router.get("/{creator_id}/lifecycle", response_model=CreatorLifecycle)
async def get_creator_lifecycle_route(
    creator_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> CreatorLifecycle:
    await _get_creator_or_404(creator_id, db, user)
    return await get_creator_lifecycle(db, creator_id)


@router.post("", response_model=CreatorOut, status_code=status.HTTP_201_CREATED)
async def create_creator(
    payload: CreatorCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Creator:
    existing = await db.execute(select(Creator).where(Creator.instagram_handle == payload.instagram_handle))
    if existing.scalar_one_or_none() is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Creator already exists")

    owner_ids = await scoped_owner_ids(user, db)
    if owner_ids is not None and payload.owner_id is not None and payload.owner_id not in owner_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Choose a valid owner within your team.")
    owner_id = payload.owner_id or user.id
    creator = Creator(
        name=payload.name,
        instagram_handle=payload.instagram_handle,
        phone=payload.phone,
        alternate_phone=payload.alternate_phone,
        email=payload.email,
        city=payload.city,
        instagram_link=payload.instagram_link or f"https://instagram.com/{payload.instagram_handle}",
        category=payload.category,
        followers_count=payload.followers_count,
        owner_id=owner_id,
        status=payload.status,
        notes=payload.notes,
    )
    db.add(creator)
    await db.flush()
    db.add(
        OwnershipEvent(
            creator_id=creator.id,
            user_id=creator.owner_id,
            event_type=OwnershipEventType.assigned,
            actor_id=user.id,
        )
    )
    await db.commit()
    await db.refresh(creator)
    return creator


@router.patch("/{creator_id}", response_model=CreatorOut)
async def update_creator(
    creator_id: int,
    payload: CreatorUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Creator:
    creator = await _get_creator_or_404(creator_id, db, user)
    update_data = payload.model_dump(exclude_unset=True)
    if user.role not in (UserRole.admin, UserRole.supervisor):
        update_data.pop("owner_id", None)
    new_owner_id = update_data.get("owner_id")
    if new_owner_id is not None and user.role == UserRole.supervisor:
        owner_ids = await scoped_owner_ids(user, db)
        if new_owner_id not in owner_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="Choose a valid owner within your team."
            )
    owner_changed = new_owner_id is not None and new_owner_id != creator.owner_id
    archive_reason = update_data.pop("archive_reason", None)
    was_archived = creator.is_archived
    for field, value in update_data.items():
        setattr(creator, field, value)
    creator.last_activity_at = datetime.now(timezone.utc)
    if creator.is_archived and not was_archived:
        if not archive_reason or not archive_reason.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="A note is required when archiving a creator.",
            )
        creator.archived_at = datetime.now(timezone.utc)
        creator.archive_reason = archive_reason
        db.add(
            OwnershipEvent(
                creator_id=creator.id,
                user_id=creator.owner_id,
                event_type=OwnershipEventType.revoked,
                actor_id=user.id,
                note=archive_reason,
            )
        )
    elif not creator.is_archived and was_archived:
        creator.archived_at = None
        creator.archive_reason = None
    if owner_changed:
        db.add(
            OwnershipEvent(
                creator_id=creator.id,
                user_id=new_owner_id,
                event_type=OwnershipEventType.admin_assigned,
                actor_id=user.id,
            )
        )
    await db.commit()
    await db.refresh(creator)
    return creator


@router.post("/{creator_id}/transfer-ownership", response_model=CreatorOut)
async def transfer_ownership(
    creator_id: int,
    payload: TransferOwnership,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Creator:
    # Self-service: the current owner (or an admin) hands this creator to
    # another advisor. _get_creator_or_404 already 404s an advisor who
    # doesn't own this creator, so no extra ownership check is needed here.
    creator = await _get_creator_or_404(creator_id, db, user)
    new_owner = await db.get(User, payload.new_owner_id)
    if new_owner is None or new_owner.role != UserRole.advisor:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Choose a valid advisor to transfer to.")
    if user.role == UserRole.supervisor:
        owner_ids = await scoped_owner_ids(user, db)
        if new_owner.id not in owner_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="Choose a valid advisor to transfer to."
            )
    if new_owner.id == creator.owner_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Creator is already owned by this user.")

    creator.owner_id = new_owner.id
    creator.last_activity_at = datetime.now(timezone.utc)
    db.add(
        OwnershipEvent(
            creator_id=creator.id,
            user_id=new_owner.id,
            event_type=OwnershipEventType.transferred,
            actor_id=user.id,
        )
    )
    await db.commit()
    await db.refresh(creator)
    return creator


@router.post("/bulk-assign", response_model=list[CreatorOut])
async def bulk_assign(
    payload: BulkAssign,
    db: AsyncSession = Depends(get_db),
    admin_user: User = Depends(require_admin),
) -> list[Creator]:
    new_owner = await db.get(User, payload.new_owner_id)
    if new_owner is None or new_owner.role != UserRole.advisor:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Choose a valid advisor to assign to.")

    result = await db.execute(select(Creator).where(Creator.id.in_(payload.creator_ids)))
    creators = list(result.scalars().all())
    now = datetime.now(timezone.utc)
    for creator in creators:
        creator.owner_id = new_owner.id
        creator.is_archived = False
        creator.archived_at = None
        creator.archive_reason = None
        creator.last_activity_at = now
        db.add(
            OwnershipEvent(
                creator_id=creator.id,
                user_id=new_owner.id,
                event_type=OwnershipEventType.admin_assigned,
                actor_id=admin_user.id,
            )
        )
    await db.commit()
    for creator in creators:
        await db.refresh(creator)
    return creators


@router.post("/{creator_id}/stage", response_model=CreatorOut)
async def transition_stage(
    creator_id: int,
    payload: StageTransition,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Creator:
    creator = await _get_creator_or_404(creator_id, db, user)

    event = StageEvent(
        creator_id=creator.id,
        from_stage=creator.current_stage,
        to_stage=payload.to_stage,
        actor_id=user.id,
        note=payload.note,
    )
    db.add(event)

    creator.current_stage = payload.to_stage
    creator.last_activity_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(creator)
    return creator


@router.get("/{creator_id}/messages", response_model=list[MessageOut])
async def list_messages(
    creator_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[MessageOut]:
    await _get_creator_or_404(creator_id, db, user)

    result = await db.execute(
        select(Message, User.name)
        .join(User, User.id == Message.author_id)
        .where(Message.creator_id == creator_id)
        .order_by(Message.created_at)
    )
    return [
        MessageOut(
            id=message.id,
            creator_id=message.creator_id,
            author_id=message.author_id,
            author_name=author_name,
            channel=message.channel,
            body=message.body,
            created_at=message.created_at,
        )
        for message, author_name in result.all()
    ]


@router.post("/{creator_id}/messages", response_model=MessageOut, status_code=status.HTTP_201_CREATED)
async def create_message(
    creator_id: int,
    payload: MessageCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> MessageOut:
    creator = await _get_creator_or_404(creator_id, db, user)

    message = Message(creator_id=creator_id, author_id=user.id, channel=payload.channel, body=payload.body)
    db.add(message)
    creator.last_activity_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(message)

    return MessageOut(
        id=message.id,
        creator_id=message.creator_id,
        author_id=message.author_id,
        author_name=user.name,
        channel=message.channel,
        body=message.body,
        created_at=message.created_at,
    )


@router.get("/{creator_id}/files", response_model=list[CreatorFileOut])
async def list_files(
    creator_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[CreatorFileOut]:
    await _get_creator_or_404(creator_id, db, user)

    result = await db.execute(
        select(CreatorFile, User.name)
        .join(User, User.id == CreatorFile.uploaded_by)
        .where(CreatorFile.creator_id == creator_id)
        .order_by(CreatorFile.created_at.desc())
    )
    return [
        CreatorFileOut(
            id=f.id,
            creator_id=f.creator_id,
            uploaded_by=f.uploaded_by,
            uploaded_by_name=uploader_name,
            original_filename=f.original_filename,
            content_type=f.content_type,
            size_bytes=f.size_bytes,
            created_at=f.created_at,
        )
        for f, uploader_name in result.all()
    ]


@router.post("/{creator_id}/files", response_model=CreatorFileOut, status_code=status.HTTP_201_CREATED)
async def upload_file(
    creator_id: int,
    upload: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> CreatorFileOut:
    await _get_creator_or_404(creator_id, db, user)

    if not upload.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File must have a name")

    contents = await upload.read()
    if len(contents) > settings.max_upload_size_bytes:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File too large")

    creator_dir = UPLOAD_ROOT / str(creator_id)
    creator_dir.mkdir(parents=True, exist_ok=True)
    stored_filename = f"{uuid.uuid4().hex}{Path(upload.filename).suffix}"
    (creator_dir / stored_filename).write_bytes(contents)

    creator_file = CreatorFile(
        creator_id=creator_id,
        uploaded_by=user.id,
        original_filename=upload.filename,
        stored_filename=stored_filename,
        content_type=upload.content_type or "application/octet-stream",
        size_bytes=len(contents),
    )
    db.add(creator_file)
    await db.commit()
    await db.refresh(creator_file)

    return CreatorFileOut(
        id=creator_file.id,
        creator_id=creator_file.creator_id,
        uploaded_by=creator_file.uploaded_by,
        uploaded_by_name=user.name,
        original_filename=creator_file.original_filename,
        content_type=creator_file.content_type,
        size_bytes=creator_file.size_bytes,
        created_at=creator_file.created_at,
    )


async def _get_creator_file_or_404(creator_id: int, file_id: int, db: AsyncSession) -> CreatorFile:
    creator_file = await db.get(CreatorFile, file_id)
    if creator_file is None or creator_file.creator_id != creator_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")
    return creator_file


@router.get("/{creator_id}/files/{file_id}/download")
async def download_file(
    creator_id: int,
    file_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> FileResponse:
    await _get_creator_or_404(creator_id, db, user)
    creator_file = await _get_creator_file_or_404(creator_id, file_id, db)

    disk_path = UPLOAD_ROOT / str(creator_id) / creator_file.stored_filename
    if not disk_path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

    return FileResponse(
        disk_path,
        media_type=creator_file.content_type,
        filename=creator_file.original_filename,
    )


@router.delete("/{creator_id}/files/{file_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_file(
    creator_id: int,
    file_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> None:
    await _get_creator_or_404(creator_id, db, user)
    creator_file = await _get_creator_file_or_404(creator_id, file_id, db)

    disk_path = UPLOAD_ROOT / str(creator_id) / creator_file.stored_filename
    disk_path.unlink(missing_ok=True)

    await db.delete(creator_file)
    await db.commit()
